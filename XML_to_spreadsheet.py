## need to not keep on computing the tree and root, create a dictionary 
## loop through once every time to create the dictionary for the root and the namespace

import xml.etree.ElementTree as ET ## allows me to pull info from xml file
import openpyxl as op
from openpyxl import load_workbook 
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import Border, Side, Alignment, Color
from openpyxl.utils import get_column_letter
import os
from helper import value, search_key, path_location_identifier, find_all, create_header_2, thin_border, create_header_1, find_namespace_declaration, fill_color_cell_grey, get_cell_number, create_row_of_black_cells



## loads open the excel file that i want to edit
wb = op.load_workbook('C:\\Users\\z004ymfp\\Downloads\\TDEF_TESTER_common1.xlsx')
sheet = wb['Common #1']

create_header_1(sheet)
## this will give me a list of all the tdef files
tdef_files = []
files_in_folder = os.listdir('TDEF_files')
for file in files_in_folder:
    # Check if the file has the ".tdef" extension
    if file.endswith('.tdef'):
        # If it does, add the file to the list of TDEF files
        tdef_files.append(file)

count = 0
headers_common1 = ['TestName', 'TestTypeID', "TestVersion", "TestName", "TestName", "TestName",
           'LOINC', 'StatusID', 'ResultReviewMode', 'ReuseResult', 'ResultTimeLimit', 'Analyte\nStability']
tdef_files_iter = iter(tdef_files)  # convert tdef_files to an iterator

## need to iterate through all the xml files and see if it is luminometer
## need to also check if it is system monitoring
filtered_tdefs = []
systemMonitoring_true = []
systemMonitoring_false = []

tdef_files_info = {} ## dictionary to carry both the namespace as well as the root of each tree 

for tdef_file in tdef_files:
    ## need to look at the detection type to see if it is a luminometer
    tree = ET.parse(f'TDEF_files/{tdef_file}')
    root = tree.getroot()
    namespace_name = find_namespace_declaration(root)
    namespace = {'ns': str(namespace_name)}
    detection_type, system_monitoring = value(tdef_file, './/ns:DetectionType', '', tree, namespace),  value(tdef_file, './/ns:IsSystemMonitoringTest', '', tree, namespace) 
    
    tdef_files_info[tdef_file] = {'tree':tree ,'root': root, 'namespace': namespace}
    
    if detection_type == 'Luminometer': ## if not luminometer, we don't consider it 
        filtered_tdefs.append(tdef_file)

        if system_monitoring == 'false':
            systemMonitoring_false.append(tdef_file)
        else:
            systemMonitoring_true.append(tdef_file)
        tdef_files_ordered = systemMonitoring_false + systemMonitoring_true
        tdef_files_ordered_iter = iter(tdef_files_ordered)

row_num = 3  # start from 3rd row
while row_num <= len(tdef_files_ordered) + 1:  # use the number of XML files
    xml_file = next(tdef_files_ordered_iter)  # get the next XML file
    
    ## access all of the information through the dictionary
    file_info = tdef_files_info[xml_file]
    tree = file_info['tree']
    root = file_info['root']
    namespace = file_info['namespace']

    row = sheet[row_num]
    if row_num  == len(systemMonitoring_false) + 3:
        create_row_of_black_cells(sheet, row_num, 1, 12, '000000')
        row_num += 1
    
    for col_num in range(1, 13):
        sheet.cell(row=row_num, column=col_num).border = thin_border

        header_value = headers_common1[count]
        

        if header_value == 'Analyte\nStability':
            value_to_add = ""
            cell = get_cell_number(row_num, col_num)
            fill_color_cell_grey(cell, sheet)
        else:
            path = search_key(path_location_identifier, header_value)
            if path:
                value_to_add = value(xml_file, path, header_value, tree, namespace)

            if header_value == "ReuseResult":
                if value_to_add == "true":
                    value_to_add = 'X'
                else:
                    value_to_add = ''
                    cell = get_cell_number(row_num, col_num)
                    fill_color_cell_grey(cell, sheet)
            if header_value == "LOINC" and value_to_add == '':
                cell = get_cell_number(row_num, col_num)
                fill_color_cell_grey(cell, sheet)
    
        # Write the value to the cell
        c = sheet.cell(row=row_num, column=col_num)
        c.value = value_to_add
    
        if count == 11:
            count = 0
            row_num += 1 
        else:
            count += 1 
    
        if row_num == len(tdef_files_ordered) + 2:
            break
        
        c = sheet.cell(row=row_num, column=col_num)
        c.value = value_to_add


new_sheet = wb.create_sheet(title='Common #2') ## create a new sheet 

## now I need to create the headers for each of these 

create_header_2(new_sheet)


## need to look at the evaluate repeatedly reactive tag/ need to look at column AH/ AD
headers_common2 = ['TestName', 'PatientReplicates', 'PatientReplicatesLock', 
                   'ControlReplicates', 'ControlReplicatesLock','PatientAcceptableCV', 
                   'PatientAcceptableCVLock', 'ControlAcceptableCV','ControlAcceptableCVLock',
                  'SpecimenID', 'ProfileID', 'CompatibilityCode', 'MinimumSoftwareVersion', 'Regions',
                  'EnterSlope', 'EnterIntercept', 'AddResult', 'IsSystemMonitoringTest', 'RecalibrateAfterUpdate',
                  'ManualDilution', 'RangeWindow', 'Anonymous', 'TestTypeID' ,'HandleRepReactive', 'AutoReDilute', 
                  'CentrifugeForRepeat', 'EnableControlBracketing', 'ControlBracketingLock', 'ControlLevels', 
                  'ControlLevelsLock' , 'TimeoutHours', 'RemindersHours', 'AdvancedDilutionOption', 
                  'AdvancedDilutionOptionLock', 'AssaySupportsADO']
count = 0
tdef_files_iter_common_2 = iter(systemMonitoring_false)
row_num = 3  # start from 3rd row

while row_num <= len(systemMonitoring_false) :  # use the number of XML files
    xml_file = next(tdef_files_iter_common_2)  # get the next XML file
    
    ## info through dictionary
    file_info = tdef_files_info[xml_file]
    tree = file_info['tree']
    root = file_info['root']
    namespace = file_info['namespace']

    row = new_sheet[row_num]
    
    for col_num in range(1, 36):
        new_sheet.cell(row=row_num, column=col_num).border = thin_border
        header_value = headers_common2[count]
        cell = get_cell_number(row_num, col_num)

        if header_value in {'PatientReplicatesLock', 'ControlReplicatesLock', 'PatientAcceptableCVLock', 'ControlAcceptableCVLock'}:
             path = search_key(path_location_identifier, header_value)
             value_to_add = value(xml_file, path, header_value, tree, namespace)
             if value_to_add == 'false':
                 value_to_add = ''
                 fill_color_cell_grey(cell, new_sheet)
             else:
                 value_to_add = 'Y'


        elif header_value in {'PatientAcceptableCV', 'ControlAcceptableCV'}:
            path = search_key(path_location_identifier, header_value)
            value_to_add = value(xml_file, path, header_value, tree, namespace)
            if value_to_add == '0.00':
                value_to_add = '0'
        
        
        
        elif header_value in {'EnterSlope', 'EnterIntercept', 'AddResult', 'IsSystemMonitoringTest', 'RecalibrateAfterUpdate',
                  'ManualDilution', 'RangeWindow', 'Anonymous', 'HandleRepReactive', 'AutoReDilute', 
                  'CentrifugeForRepeat', 'EnableControlBracketing', 'ControlBracketingLock', 'ControlLevelsLock', 'AdvancedDilutionOption', 
                   'AssaySupportsADO'}:
             path = search_key(path_location_identifier, header_value)
             value_to_add = value(xml_file, path, header_value, tree, namespace)
             if value_to_add == 'true':
                value_to_add = 'X'
             else:
                value_to_add = ''
                cell = get_cell_number(row_num, col_num)
                fill_color_cell_grey(cell, new_sheet)
        elif header_value == 'TestTypeID':
            
            path = search_key(path_location_identifier, header_value)
            value_to_add = value(xml_file, path, header_value, tree, namespace)
            if value_to_add == 'IDAutoFRR':
                value_to_add = 'X'
            else:
                value_to_add = ''
                cell = get_cell_number(row_num, col_num)
                fill_color_cell_grey(cell, new_sheet)
        
        elif header_value in {'ControlLevels' , 'TimeoutHours', 'RemindersHours'}:
            path = search_key(path_location_identifier, header_value)
            value_to_add = value(xml_file, path, header_value, tree, namespace)
            if value_to_add == '0':
                value_to_add = ''
                cell = get_cell_number(row_num, col_num)
                fill_color_cell_grey(cell, new_sheet)
        
        
        elif header_value in {'SpecimenID', 'ProfileID'}:  
            items = find_all(root, namespace, header_value)
            value_to_add = '\n'.join(items) if items else ''
            cell = new_sheet.cell(row=row_num, column=col_num)
            cell.value = value_to_add
            cell.alignment = cell.alignment.copy(wrap_text=True, horizontal='center', vertical='center')
            
        elif header_value == 'AdvancedDilutionOptionLock':
            value_to_add = ''
            cell = get_cell_number(row_num, col_num)
            fill_color_cell_grey(cell, new_sheet)
            
        else:
            path = search_key(path_location_identifier, header_value)
            if path:
                value_to_add = value(xml_file, path, header_value, tree, namespace)                 
                
        if value_to_add == '0.00':
            value_to_add == '0'
        if value_to_add == '1.0':
            value_to_add = '1'
 
        # Write the value to the cell
        c = new_sheet.cell(row=row_num, column=col_num)
        c.value = value_to_add
        c.alignment = c.alignment.copy(wrap_text=True, horizontal='center', vertical='center')


        if count == 34:
            count = 0
            row_num += 1 
        else:
            count += 1 
    
        if row_num == len(tdef_files_ordered) + 2:
            break
        c = new_sheet.cell(row=row_num, column=col_num)
        c.value = value_to_add
      
wb.save('C:\\Users\\z004ymfp\\Downloads\\TDEF_TESTER_common1.xlsx')











